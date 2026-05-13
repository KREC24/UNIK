"""
Сквозной скрипт: парсинг ведомостей ОГЗ + расчёт ПТМ + генерация Excel.

Вход: testdoc1 (два PDF: R45 и R60)
Выход: testxl1/МФК_Адапт_Спорт_R45.xlsx и testxl1/МФК_Адапт_Спорт_R60.xlsx
"""

import os, re, sys
from pathlib import Path
from decimal import Decimal
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent.parent.parent / "backend"))

PROJECT_ROOT = Path(__file__).parent.parent.parent
TESTDOC1 = PROJECT_ROOT / "testdoc1_extracted"
TESTXL1 = PROJECT_ROOT / "testxl1"
TESTXL1.mkdir(parents=True, exist_ok=True)

STEEL_DENSITY = 7850.0

COMPOSITIONS = {
    "R45": {
        "name": "Таггерт Экстрим (орг. основа)",
        "grunt": "Грунт ГФ-021 (зимний)",
        "kraska": "Таггерт Экстрим R45 (органорастворимый)",
        "finish": "Финиш ПФ-115",
        "kraska_rate_kg_m2_mm": 0.58,
    },
    "R60": {
        "name": "Таггерт Экстрим (орг. основа, усиленный)",
        "grunt": "Грунт ГФ-021 (зимний)",
        "kraska": "Таггерт Экстрим R60 (органорастворимый)",
        "finish": "Финиш ПФ-115",
        "kraska_rate_kg_m2_mm": 0.88,
    },
    "R120": {
        "name": "Таггерт К-Экстрим (конструктивная ОГЗ)",
        "grunt": "Грунт ГФ-021 (зимний)",
        "kraska": "Таггерт К-Экстрим (ТИ+ОЗП)",
        "finish": "Финиш ПФ-115",
        "kraska_rate_kg_m2_mm": 1.16,
    },
}

PROJECT_META = {
    "name": "МФК по адаптивным видам спорта",
    "address": "г. Красноярск, пер. Афонтовский, 7",
    "stage": "0 цикл",
}


def find_pdfs(base_dir: Path, keyword: str) -> list[Path]:
    result = []
    for root, dirs, files in os.walk(base_dir):
        for f in files:
            if f.endswith(".pdf") and keyword.lower() in f.lower():
                result.append(Path(root) / f)
    return sorted(result)


def extract_data_from_pdf(pdf_path: Path) -> list[dict]:
    """Извлекает строки данных из PDF ведомости ОГЗ."""
    items = []

    with pdfplumber.open(pdf_path) as pdf:
        full_text = "\n".join(
            page.extract_text() or "" for page in pdf.pages
        )

    lines = full_text.split("\n")
    current_section = ""

    # Паттерн для строки данных:
    # Профиль | Масса, т | Площадь, м2 | Покрытие | R | ПТМ | ТИ мм | ТИ кг/м2 | ОЗП мм | ОЗП кг/м2 | % | ТИ всего | ОЗП всего
    # Пример: ☐ 100х4 0,34 12,17 К-Экстрим R120 3,87 1,74 2,26 1,16 1,62 25% 35,78 25,69
    # Пример: I 30К1 0,68 15,03 Таггерт Экстрим R120 6,35 一 一 1,24 1,91 25% 一 37,41

    data_pattern = re.compile(
        r"^\s*([\u2610\w\s\d\.\-×xх]+?)\s+"       # profile (with ☐)
        r"([\d,]+\s*[\d,]*)\s+"                      # mass, tons
        r"([\d,]+\s*[\d,]*)\s+"                      # area, m2
        r"(\S+(?:\s+\S+)*?)\s+"                      # coating
        r"R(\d+)\s+"                                   # R value
        r"([\d,]+)\s+"                                 # PTM, mm
        r"(\S+)\s+"                                     # TI layer mm (может быть 一 или —)
        r"(\S+)\s+"                                     # TI rate kg/m2
        r"(\S+)\s+"                                     # OZP layer mm
        r"(\S+)\s+"                                     # OZP rate kg/m2
        r"(\d+%)\s+"                                    # loss %
        r"(\S+)\s+"                                     # TI total kg
        r"([\d,]+\s*[\d,]*)\s*$"                       # OZP total kg
    )

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # Секции (надколоннки, связи, фермы, балки...)
        if line.endswith("и") and len(line) < 40 and not any(c.isdigit() for c in line):
            current_section = line.rstrip("и").strip()
            continue

        m = data_pattern.match(line)
        if m:
            profile = m.group(1).strip()
            mass_str = m.group(2).replace(",", ".").replace(" ", "")
            area_str = m.group(3).replace(",", ".").replace(" ", "")
            coating = m.group(4).strip()
            r_value = int(m.group(5))
            ptm_str = m.group(6).replace(",", ".")
            ti_mm_str = m.group(7)
            ti_rate_str = m.group(8)
            ozp_mm_str = m.group(9)
            ozp_rate_str = m.group(10)
            loss_str = m.group(11)
            ti_total_str = m.group(12)
            ozp_total_str = m.group(13).replace(",", ".").replace(" ", "")

            try:
                mass = float(mass_str)
                area = float(area_str)
                ptm = float(ptm_str)
                ozp_total = float(ozp_total_str)
            except (ValueError, IndexError):
                continue

            items.append({
                "section": current_section,
                "profile": profile,
                "mass_ton": mass,
                "area_m2": area,
                "coating": coating,
                "r_value": r_value,
                "ptm_mm": ptm,
                "ti_layer_mm": ti_mm_str,
                "ti_rate_kg_m2": ti_rate_str,
                "ozp_layer_mm": ozp_mm_str,
                "ozp_rate_kg_m2": ozp_rate_str,
                "loss_pct": loss_str,
                "ti_total_kg": ti_total_str,
                "ozp_total_kg": ozp_total,
            })

    return items


def calculate_ogz(items: list[dict], target_r: int) -> list[dict]:
    """Рассчитывает состав ОГЗ для каждого элемента под заданный R."""
    result = []

    for item in items:
        ptm = item["ptm_mm"]
        area = item["area_m2"]

        # Выбор состава по R
        comp_key = f"R{target_r}" if f"R{target_r}" in COMPOSITIONS else (
            "R45" if target_r <= 45 else "R60" if target_r <= 60 else "R120"
        )
        comp = COMPOSITIONS.get(comp_key, COMPOSITIONS["R60"])

        # Расход краски = площадь (м2) × ПТМ (мм) × норма (кг/м2/мм)
        kraska_rate = comp["kraska_rate_kg_m2_mm"]
        kraska_kg = area * ptm * kraska_rate

        # Грунт: 0.35 кг/м2 (зимний, влажная погода, осень)
        grunt_kg = area * 0.35

        # Финиш: 0.25 кг/м2
        finish_kg = area * 0.25

        item["recommended_ogz"] = comp["name"]
        item["kraska_consumption_kg"] = round(kraska_kg, 2)
        item["grunt_consumption_kg"] = round(grunt_kg, 2)
        item["finish_consumption_kg"] = round(finish_kg, 2)
        result.append(item)

    return result


def generate_excel(items: list[dict], output_path: Path, r_label: str):
    """Генерирует Excel-файл, стилизованный под эталон из testxl."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Ведомость ОГЗ R{r_label}"

    # Заголовки
    headers = [
        "№ п/п", "Секция", "Профиль", "Масса металла, т",
        "Защищаемая площадь, м²", "Предел R, мин",
        "ПТМ, мм", "Рекомендуемый состав ОГЗ",
        "Расход грунта, кг", "Расход краски, кг",
        "Расход финиша, кг", "ОЗП всего (эталон), кг",
    ]

    # Стили
    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Calibri", size=10)
    cell_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    total_fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")

    # Строка метаданных
    ws.merge_cells("A1:L1")
    ws["A1"] = f"Объект: {PROJECT_META['name']} | Адрес: {PROJECT_META['address']} | Стадия: {PROJECT_META['stage']} | Расчёт: R{r_label}"
    ws["A1"].font = Font(name="Calibri", bold=True, size=13, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Пустая строка
    ws.merge_cells("A3:L3")
    ws["A3"] = f"Приложение к КП по огнезащите металлоконструкций. Основа: органическая (осенне-зимний период)."
    ws["A3"].font = Font(name="Calibri", italic=True, size=9, color="666666")

    # Заголовки таблицы
    header_row = 5
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Данные
    data_start = header_row + 1
    total_mass = 0.0
    total_area = 0.0
    total_kraska = 0.0
    total_grunt = 0.0
    total_finish = 0.0
    total_ozp_ref = 0.0

    for i, item in enumerate(items):
        row = data_start + i
        total_mass += item["mass_ton"]
        total_area += item["area_m2"]
        total_kraska += item.get("kraska_consumption_kg", 0)
        total_grunt += item.get("grunt_consumption_kg", 0)
        total_finish += item.get("finish_consumption_kg", 0)
        total_ozp_ref += item["ozp_total_kg"]

        values = [
            i + 1,
            item.get("section", ""),
            item["profile"],
            item["mass_ton"],
            item["area_m2"],
            item["r_value"],
            item["ptm_mm"],
            item.get("recommended_ogz", ""),
            item.get("grunt_consumption_kg", 0),
            item.get("kraska_consumption_kg", 0),
            item.get("finish_consumption_kg", 0),
            item["ozp_total_kg"],
        ]

        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col_idx, value=val)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border

    # Итоги
    total_row = data_start + len(items)
    totals = [
        "", "", "ИТОГО:", round(total_mass, 2), round(total_area, 2), "",
        "", "", round(total_grunt, 2), round(total_kraska, 2),
        round(total_finish, 2), round(total_ozp_ref, 2),
    ]
    for col_idx, val in enumerate(totals, 1):
        cell = ws.cell(row=total_row, column=col_idx, value=val)
        cell.font = Font(name="Calibri", bold=True, size=10)
        cell.fill = total_fill
        cell.alignment = cell_align
        cell.border = thin_border

    # Ширина колонок
    col_widths = [7, 14, 18, 14, 18, 10, 8, 30, 14, 14, 14, 18]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Высота строк
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[header_row].height = 35

    # Автофильтр
    ws.auto_filter.ref = f"A{header_row}:L{total_row}"

    wb.save(output_path)
    print(f"  Сохранён: {output_path}")
    return {"count": len(items), "total_mass": total_mass, "total_area": total_area, "total_kraska": total_kraska}


def main():
    print("=" * 70)
    print("Парсинг ведомостей ОГЗ для МФК Адаптивные виды спорта")
    print("=" * 70)

    # Ищем PDF ведомостей
    r45_pdfs = find_pdfs(TESTDOC1, "R45")
    r60_pdfs = find_pdfs(TESTDOC1, "R60")

    if not r45_pdfs:
        print("ERROR: PDF с R45 не найден")
        return
    if not r60_pdfs:
        print("ERROR: PDF с R60 не найден")
        return

    print(f"R45 PDF: {r45_pdfs[0].name}")
    print(f"R60 PDF: {r60_pdfs[0].name}")

    # Парсим оба PDF
    print("\n--- Парсинг R45 ---")
    items_r45 = extract_data_from_pdf(r45_pdfs[0])
    print(f"  Извлечено строк: {len(items_r45)}")

    print("\n--- Парсинг R60 ---")
    items_r60 = extract_data_from_pdf(r60_pdfs[0])
    print(f"  Извлечено строк: {len(items_r60)}")

    # Рассчитываем ОГЗ
    print("\n--- Расчёт ОГЗ ---")
    items_r45 = calculate_ogz(items_r45, 45)
    items_r60 = calculate_ogz(items_r60, 60)
    print(f"  R45: {len(items_r45)} позиций рассчитано")
    print(f"  R60: {len(items_r60)} позиций рассчитано")

    # Генерируем Excel
    print("\n--- Генерация Excel ---")
    stats_r45 = generate_excel(items_r45, TESTXL1 / "МФК_Адапт_Спорт_R45.xlsx", "45")
    stats_r60 = generate_excel(items_r60, TESTXL1 / "МФК_Адапт_Спорт_R60.xlsx", "60")

    print(f"\n{'='*70}")
    print(f"RESULT:")
    print(f"  R45: {stats_r45['count']} rows | mass {stats_r45['total_mass']:.2f} t | area {stats_r45['total_area']:.2f} m2 | paint {stats_r45['total_kraska']:.0f} kg")
    print(f"  R60: {stats_r60['count']} rows | mass {stats_r60['total_mass']:.2f} t | area {stats_r60['total_area']:.2f} m2 | paint {stats_r60['total_kraska']:.0f} kg")
    print(f"  Files: {TESTXL1}")
    print(f"{'='*70}")


if __name__ == "__main__":
    main()
