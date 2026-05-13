"""Финал: парсинг testdoc + сверка + Excel."""

import sys, os, json, re
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent.parent))
from app.core.kmd_parser import KmdShippingParser
from app.services.parsing_service import parse_revC04

ROOT = Path(__file__).parent.parent.parent
TESTDOC = ROOT / "testdoc"
TESTXL = ROOT / "testxl"
REF = TESTXL / "Shiping_list_выбр_18.04.25 по 5.2 ведомость.xls"
OUT = TESTXL / "Ведомость_КМД_ПАРСИНГ.xlsx"

# --- Эталон ---
df = pd.read_excel(REF, header=None)
ref = {}
for _, row in df.iterrows():
    m = str(row[1]).strip() if pd.notna(row[1]) else ""
    if not m or m == "nan":
        continue

    def sf(v):
        if pd.isna(v):
            return None
        s = str(v).replace(",", ".").replace(" ", "")
        try:
            return float(s)
        except:
            return None

    w = sf(row[9])
    a = sf(row[11])
    if w is None and a is None:
        continue

    ref[m] = {
        "type_name": str(row[3]).strip() if pd.notna(row[3]) else None,
        "quantity": int(sf(row[4])) if sf(row[4]) else None,
        "length_x": sf(row[5]), "width_y": sf(row[6]), "height_z": sf(row[7]),
        "unit_weight_kg": sf(row[8]), "total_weight_kg": w,
        "unit_area_m2": sf(row[10]), "total_area_m2": a,
    }

REF_MASS = 263812.60
REF_AREA = 5623.44

# --- Парсинг ---
print("=" * 60)
print("UNIK - Финальный парсинг testdoc")
print("=" * 60)

items = []

# л.2
l2 = list(TESTDOC.glob("*л.2*"))
if l2:
    p = KmdShippingParser(l2[0])
    r = p.parse()
    print(f"л.2: {r.total_rows_parsed}/{r.total_rows_raw} строк")
    for item in r.items:
        items.append({
            "mark": item.get("mark"),
            "type_name": item.get("type_name"),
            "quantity": float(item["quantity"]) if item.get("quantity") else None,
            "length_x": float(item["length_x"]) if item.get("length_x") else None,
            "width_y": float(item["width_y"]) if item.get("width_y") else None,
            "height_z": float(item["height_z"]) if item.get("height_z") else None,
            "unit_weight_kg": float(item["unit_weight_kg"]) if item.get("unit_weight_kg") else None,
            "total_weight_kg": float(item["total_weight_kg"]) if item.get("total_weight_kg") else None,
            "unit_area_m2": float(item["unit_area_m2"]) if item.get("unit_area_m2") else None,
            "total_area_m2": float(item["total_area_m2"]) if item.get("total_area_m2") else None,
        })

# revC04
rc4 = list(TESTDOC.glob("*revC04*"))
if rc4:
    r2 = parse_revC04(rc4[0])
    print(f"revC04: {r2.total_rows_parsed}/{r2.total_rows_raw} строк")

# Дедупликация (только items с mass)
seen = {}
for it in items:
    mk = it["mark"]
    if not mk:
        continue
    if mk not in seen or (it.get("total_weight_kg") or 0) > (seen[mk].get("total_weight_kg") or 0):
        seen[mk] = it

merged = list(seen.values())
merged.sort(key=lambda x: x.get("total_weight_kg") or 0, reverse=True)

our_mass = sum(i.get("total_weight_kg") or 0 for i in merged)
our_area = sum(i.get("total_area_m2") or 0 for i in merged)
our_marks = {i["mark"] for i in merged}
ref_marks = set(ref.keys())

missing = ref_marks - our_marks
miss_mass = sum(ref[m].get("total_weight_kg") or 0 for m in missing)
miss_area = sum(ref[m].get("total_area_m2") or 0 for m in missing)

print(f"\nЭталон: {REF_MASS:.0f} кг, {REF_AREA:.2f} m2, {len(ref_marks)} позиций")
print(f"Парсер: {our_mass:.0f} кг, {our_area:.2f} m2, {len(our_marks)} позиций")
print(f"Покрытие: {our_mass/REF_MASS*100:.1f}% массы, {our_area/REF_AREA*100:.1f}% площади")
print(f"Не хватает: {len(missing)} позиций, {miss_mass:.0f} кг, {miss_area:.2f} m2")

# --- Excel ---
wb = Workbook()
ws = wb.active
ws.title = "Ведомость КМД"

h = ["Поз", "Марка", "", "Описание", "Кол-во", "X", "Y", "Z", "Масса ед.", "Масса общ.", "S ед.", "S общ."]
hf = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
hfl = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
bd = Border(left=Side("thin"), right=Side("thin"), top=Side("thin"), bottom=Side("thin"))
ha = Alignment(horizontal="center", vertical="center")

ws.merge_cells("A1:L1")
ws["A1"] = f"Ведомость отправочных марок | Проект: В068522/1540Д-416-55500-01-5.2-КМД | Авто-парсинг | Покрытие: {our_mass/REF_MASS*100:.1f}%"
ws["A1"].font = Font(name="Calibri", bold=True, size=12, color="1F4E79")
ws["A1"].alignment = Alignment(horizontal="center")

for ci, hdr in enumerate(h, 1):
    c = ws.cell(row=3, column=ci, value=hdr)
    c.font = hf; c.fill = hfl; c.alignment = ha; c.border = bd

for i, item in enumerate(merged):
    row = 4 + i
    v = [i+1, item.get("mark"), "", item.get("type_name"), item.get("quantity"),
         item.get("length_x"), item.get("width_y"), item.get("height_z"),
         item.get("unit_weight_kg"), item.get("total_weight_kg"),
         item.get("unit_area_m2"), item.get("total_area_m2")]
    for ci, val in enumerate(v, 1):
        c = ws.cell(row=row, column=ci, value=val)
        c.font = Font(name="Calibri", size=10); c.alignment = ha; c.border = bd

tr = 4 + len(merged)
tfl = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
tv = ["ИТОГО:", "", "", "", "", "", "", "", "", round(our_mass, 2), "", round(our_area, 2)]
for ci, val in enumerate(tv, 1):
    c = ws.cell(row=tr, column=ci, value=val)
    c.font = Font(name="Calibri", bold=True, size=10); c.fill = tfl; c.alignment = ha; c.border = bd

for i, w in enumerate([7, 14, 3, 18, 8, 9, 9, 9, 13, 14, 11, 12], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

ws.row_dimensions[1].height = 25
ws.auto_filter.ref = f"A3:L{tr}"

wb.save(OUT)
print(f"\nФайл: {OUT}")
print(f"Строк: {len(merged)}")
