"""
Парсер КМД-данных из проектной документации testdoc1.
Извлекает чистую ведомость металла → LineItem (без расчёта ОГЗ).

Формат выхода: Поз, Марка, Описание, Кол-во, Габариты X×Y×Z,
Масса ед/общ, Площадь ед/общ, ПТМ, Профиль, Марка стали.
"""

import os, re, sys
from pathlib import Path
from decimal import Decimal
import pdfplumber
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent.parent.parent / "backend"))

PROJECT_ROOT = Path(__file__).parent.parent.parent
TESTDOC1 = PROJECT_ROOT / "testdoc1_extracted"
TESTXL1 = PROJECT_ROOT / "testxl1"
TESTXL1.mkdir(parents=True, exist_ok=True)

PROJECT_META = {
    "name": "МФК по адаптивным видам спорта",
    "address": "г. Красноярск, пер. Афонтовский, 7",
    "stage": "0 цикл",
    "project_code": "149-24-ПБ",
}


def find_pdfs(base_dir: Path, keyword: str) -> list[Path]:
    result = []
    for root, dirs, files in os.walk(base_dir):
        for f in files:
            if f.endswith(".pdf") and keyword.lower() in f.lower():
                result.append(Path(root) / f)
    return sorted(result)


def extract_line_items_from_pdf(pdf_path: Path) -> list[dict]:
    """
    Извлекает строки КМД из PDF.
    Маппинг на LineItem: mark, type_name, quantity, dims, mass, area, ptm.
    """
    items = []

    with pdfplumber.open(pdf_path) as pdf:
        full_text = "\n".join(
            page.extract_text() or "" for page in pdf.pages
        )

    lines = full_text.split("\n")
    current_section = ""

    # Паттерн: Профиль | Масса, т | Площадь, м2 | Покрытие | R | ПТМ | ...
    data_pattern = re.compile(
        r"^\s*([\u2610\w\s\d\.\-×xXхHØ]+?)\s+"
        r"([\d,]+[\s\d,]*)\s+"
        r"([\d,]+[\s\d,]*)\s+"
        r"(\S.*?\S)\s+"
        r"R(\d+)\s+"
        r"([\d,]+)"
    )

    for line in lines:
        line = line.strip()
        if not line:
            continue

        # Заголовки секций
        section_keywords = [
            "Надколоннк", "Связ", "Ферм", "Балк", "Прогон",
            "Колонн", "Стойк", "Распорк", "Ригел", "Обвяз",
        ]
        for kw in section_keywords:
            if line.startswith(kw) and len(line) < 40 and not any(c.isdigit() for c in line):
                current_section = line.rstrip("и").strip()
                break

        m = data_pattern.match(line)
        if not m:
            continue

        profile = m.group(1).strip()
        mass_str = m.group(2).replace(",", ".").replace(" ", "")
        area_str = m.group(3).replace(",", ".").replace(" ", "")
        coating = m.group(4).strip()
        r_value = int(m.group(5))
        ptm_str = m.group(6).replace(",", ".")

        try:
            mass_ton = float(mass_str)
            area_m2 = float(area_str)
            ptm_mm = float(ptm_str)
        except (ValueError, IndexError):
            continue

        total_weight_kg = round(mass_ton * 1000, 1)
        total_area_m2 = round(area_m2, 2)

        items.append({
            "position": None,
            "mark": profile,
            "type_name": current_section or _classify_element(profile, current_section),
            "quantity": None,
            "length_x": None,
            "width_y": None,
            "height_z": None,
            "unit_weight_kg": None,
            "total_weight_kg": Decimal(str(total_weight_kg)),
            "unit_area_m2": None,
            "total_area_m2": Decimal(str(total_area_m2)),
            "ptm": Decimal(str(ptm_mm)),
            "ogz_notes": f"R{r_value} | {coating}",
            "profile_type": _classify_profile_type(profile),
            "steel_grade": None,
            "source_sheet": os.path.basename(pdf_path),
        })

    return items


def _classify_element(profile: str, section: str) -> str:
    """Определяет тип элемента по профилю и секции."""
    section_lower = section.lower()
    if "надколон" in section_lower:
        return "Надколонник"
    if "связ" in section_lower:
        return "Связь"
    if "ферм" in section_lower:
        return "Ферма"
    if "балк" in section_lower:
        return "Балка"
    if "прогон" in section_lower:
        return "Прогон"
    if "колон" in section_lower:
        return "Колонна"
    if "стойк" in section_lower:
        return "Стойка"
    if "распор" in section_lower:
        return "Распорка"
    if "ригел" in section_lower:
        return "Ригель"
    if "обвяз" in section_lower:
        return "Обвязка"

    # Определение по профилю
    p = profile.upper()
    if p.startswith("I"):
        return "Балка"
    if p.startswith("\u2610") or "П" in p:
        return "Труба профильная"
    if p.startswith("L"):
        return "Уголок"
    if p.startswith("Ø"):
        return "Круг"
    return section if section else "Элемент"


def _classify_profile_type(profile: str) -> str:
    p = profile.upper().strip()
    if p.startswith("I"):
        return "beam"
    if p.startswith("\u2610") or "[" in p or "П" in p:
        return "pipe_box"
    if p.startswith("L"):
        return "angle"
    if p.startswith("Ø"):
        return "round"
    if "X" in p or "Х" in p:
        return "pipe"
    return "other"


def generate_sheet_excel(items: list[dict], output_path: Path, sheet_label: str):
    """Генерирует Excel в формате ведомости КМД (совместим с Shiping_list)."""
    wb = Workbook()
    ws = wb.active
    ws.title = f"Ведомость КМД ({sheet_label})"

    headers = [
        "Поз", "Марка (профиль)", "Описание", "Кол-во",
        "X, мм", "Y, мм", "Z, мм",
        "Масса ед., кг", "Масса общ., кг",
        "S ед., м²", "S общ., м²",
        "ПТМ, мм", "Примечания",
    ]

    header_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Calibri", size=10)
    cell_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    total_fill = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

    # Метаданные
    ws.merge_cells("A1:M1")
    ws["A1"] = (
        f"Сводная ведомость КМД | Объект: {PROJECT_META['name']} | "
        f"Адрес: {PROJECT_META['address']} | Стадия: {PROJECT_META['stage']} | "
        f"Шифр: {PROJECT_META['project_code']}"
    )
    ws["A1"].font = Font(name="Calibri", bold=True, size=12, color="1F4E79")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Источник данных
    ws.merge_cells("A3:M3")
    ws["A3"] = f"Источник: проектная документация КМ/КМД. Парсинг: автоматический. Расчёт ОГЗ не выполнен (чистый металл)."
    ws["A3"].font = Font(name="Calibri", italic=True, size=9, color="808080")

    # Заголовки
    hr = 5
    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=hr, column=ci, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # Данные
    dr = hr + 1
    totals = {"mass": 0.0, "area": 0.0, "count": 0}

    for i, item in enumerate(items):
        row = dr + i
        tw = float(item["total_weight_kg"] or 0)
        ta = float(item["total_area_m2"] or 0)
        ptm = float(item["ptm"] or 0)

        totals["mass"] += tw
        totals["area"] += ta
        totals["count"] += 1

        vals = [
            i + 1,
            item["mark"],
            item["type_name"],
            item["quantity"],
            item["length_x"],
            item["width_y"],
            item["height_z"],
            item["unit_weight_kg"],
            tw,
            item["unit_area_m2"],
            ta,
            ptm,
            item.get("ogz_notes", ""),
        ]

        for ci, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=ci, value=v)
            cell.font = cell_font
            cell.alignment = cell_align
            cell.border = thin_border

    # Итоги
    tr = dr + len(items)
    total_vals = [
        "", "", "ИТОГО:", "",
        "", "", "", "",
        round(totals["mass"], 1),
        "",
        round(totals["area"], 2),
        "", "",
    ]
    for ci, v in enumerate(total_vals, 1):
        cell = ws.cell(row=tr, column=ci, value=v)
        cell.font = Font(name="Calibri", bold=True, size=10)
        cell.fill = total_fill
        cell.alignment = cell_align
        cell.border = thin_border

    # Ширина колонок
    widths = [6, 18, 18, 8, 9, 9, 9, 13, 14, 11, 12, 9, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[hr].height = 32
    ws.auto_filter.ref = f"A{hr}:M{tr}"

    wb.save(output_path)
    return totals


def main():
    print("=" * 70)
    print("Парсинг КМД: МФК Адаптивные виды спорта |> чистая ведомость металла")
    print("=" * 70)

    r45_pdfs = find_pdfs(TESTDOC1, "R45")
    r60_pdfs = find_pdfs(TESTDOC1, "R60")

    items_r45 = extract_line_items_from_pdf(r45_pdfs[0]) if r45_pdfs else []
    items_r60 = extract_line_items_from_pdf(r60_pdfs[0]) if r60_pdfs else []

    # Объединяем уникальные позиции (по профилю + площади)
    all_items = {}
    for item in items_r45 + items_r60:
        key = f"{item['mark']}|{item['total_area_m2']}"
        if key not in all_items:
            all_items[key] = item

    merged = list(all_items.values())
    # Сортируем по массе (крупные элементы в начале)
    merged.sort(key=lambda x: float(x["total_weight_kg"] or 0), reverse=True)

    print(f"R45: {len(items_r45)} строк")
    print(f"R60: {len(items_r60)} строк")
    print(f"Уникальных позиций: {len(merged)}")

    # Генерируем единую ведомость КМД
    output = TESTXL1 / "Сводная_ведомость_КМД_МФК.xlsx"
    totals = generate_sheet_excel(merged, output, "сводная")

    print(f"\n{'='*70}")
    print(f"Сводная ведомость КМД:")
    print(f"  Позиций: {totals['count']}")
    print(f"  Масса общая: {totals['mass']:.1f} кг ({totals['mass']/1000:.2f} т)")
    print(f"  Площадь общая: {totals['area']:.2f} m2")
    print(f"  Файл: {output}")
    print(f"{'='*70}")


if __name__ == "__main__":
    main()
